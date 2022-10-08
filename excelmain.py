import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["products"]

productsPerSupplier = {}
pricePerCompany = {}
lessThan10 = {}

for i in range(2, product_list.max_row+1):
    supplierName = product_list.cell(i, 4).value
    inventoryValue = product_list.cell(i, 2).value
    productPrice = product_list.cell(i, 3).value
    productNumber = product_list.cell(i,1).value

    multiply = inventoryValue * productPrice
    #
    if supplierName not in productsPerSupplier:
        productsPerSupplier[supplierName] = 1
    else:
        productsPerSupplier[supplierName] = productsPerSupplier[supplierName] + 1

    if supplierName not in pricePerCompany:
        pricePerCompany[supplierName] = multiply
    else:
        pricePerCompany[supplierName] = pricePerCompany.get(supplierName) + multiply


print(productsPerSupplier)
print(pricePerCompany)

print(type(productsPerSupplier.get("AAA Company")))
